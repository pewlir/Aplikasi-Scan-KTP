import streamlit as st
from PIL import Image
import pytesseract
import pandas as pd
import os, re, cv2, numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# === KONFIGURASI APLIKASI ===
st.set_page_config(page_title="Smart OCR KTP / SIM / NPWP", page_icon="🪪", layout="wide")
st.title("🧠 Smart OCR – Deteksi & Validasi Otomatis (KTP / SIM / NPWP)")

EXCEL_FILE = "data_kartu.xlsx"

# === PERSIAPAN DATA ===
def load_data():
    if os.path.exists(EXCEL_FILE):
        return pd.read_excel(EXCEL_FILE)
    else:
        df = pd.DataFrame(columns=[
            "Batch ID", "Waktu", "Nama File", "Jenis Dokumen",
            "NIK / Nomor", "Nama", "Tempat/Tanggal Lahir", "Alamat",
            "Keterangan Tambahan", "Teks Asli"
        ])
        df.to_excel(EXCEL_FILE, index=False)
        return df

def save_data(df):
    df.to_excel(EXCEL_FILE, index=False)

def append_data(new_rows):
    df = load_data()
    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    save_data(df)

# === IMAGE UTILITIES ===
def auto_rotate_and_crop(image):
    img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blur, 50, 150)
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if contours:
        c = max(contours, key=cv2.contourArea)
        rect = cv2.minAreaRect(c)
        angle = rect[-1]
        if angle < -45:
            angle = 90 + angle
        (h, w) = img_cv.shape[:2]
        M = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
        rotated = cv2.warpAffine(img_cv, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        gray_rot = cv2.cvtColor(rotated, cv2.COLOR_BGR2GRAY)
        edges_rot = cv2.Canny(gray_rot, 50, 150)
        contours2, _ = cv2.findContours(edges_rot, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if contours2:
            c2 = max(contours2, key=cv2.contourArea)
            x, y, w, h = cv2.boundingRect(c2)
            cropped = rotated[y:y+h, x:x+w]
            return Image.fromarray(cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB))
    return image

def preprocess_image(image):
    img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    gray = cv2.medianBlur(gray, 3)
    gray = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                 cv2.THRESH_BINARY, 31, 5)
    resized = cv2.resize(gray, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
    return Image.fromarray(resized)

# === DETEKSI JENIS DOKUMEN ===
def detect_document_type(text):
    t = text.upper()
    if any(word in t for word in ["KARTU TANDA PENDUDUK", "NIK", "PROVINSI", "KECAMATAN"]):
        return "KTP"
    elif any(word in t for word in ["SURAT IZIN MENGEMUDI", "POLRI", "GOL.", "NO SIM"]):
        return "SIM"
    elif any(word in t for word in ["NPWP", "DIREKTORAT JENDERAL PAJAK", "PAJAK"]):
        return "NPWP"
    else:
        return "Tidak Dikenal"

# === SMART CORRECTION ===
def smart_correct_text(text):
    corrections = {
        "N1K": "NIK", "NIKK": "NIK", "1K": "IK",
        "PEKERJAAAN": "PEKERJAAN", "KEL/": "KELURAHAN ",
        "KEL.": "KELURAHAN ", "KECAMATN": "KECAMATAN",
        "TEMPAT/TGI LAHIR": "TEMPAT/TGL LAHIR"
    }
    for k, v in corrections.items():
        text = text.replace(k, v)
    return text

# === PARSING ===
def parse_document_fields(text, doc_type):
    text = smart_correct_text(text.upper())

    if doc_type == "KTP":
        # Bersihkan noise
        text = text.replace("|", "I").replace(";", ":").replace("‘", "'")
        text = re.sub(r"[^A-Z0-9\s:.,/-]", " ", text)

        nik = re.search(r"NIK[:\s]*([\d]{13,17})", text)
        nama = re.search(r"NAMA[:\s]*([A-Z\s]+)", text)
        ttl = re.search(r"(TEMPAT|TMPT)[/\s]*(TGL|TANGGAL)*\s*LAHIR[:\s]*([A-Z,\s0-9-]+)", text)
        jk = re.search(r"JENIS\s*KELAMIN[:\s]*([A-Z\s]+)", text)
        gol_darah = re.search(r"GOL[:\s]*([ABO]{1,2}\+?)", text)
        alamat = re.search(r"ALAMAT[:\s]*([A-Z0-9\s.,/-]+)", text)
        rt_rw = re.search(r"RT/RW[:\s]*([\d\s/]+)", text)
        kel_desa = re.search(r"(KEL|DESA)[:\s]*([A-Z\s]+)", text)
        kecamatan = re.search(r"KEC(AMATAN)*[:\s]*([A-Z\s]+)", text)
        agama = re.search(r"AGAMA[:\s]*([A-Z]+)", text)
        status = re.search(r"STATUS[:\s]*(PERKAWINAN)*[:\s]*([A-Z\s]+)", text)
        pekerjaan = re.search(r"PEKERJAAN[:\s]*([A-Z\s]+)", text)
        kewarganegaraan = re.search(r"KEWARGANEGARAAN[:\s]*([A-Z]+)", text)
        berlaku = re.search(r"BERLAKU[:\s]*HINGGA[:\s]*([A-Z0-9\s-]+)", text)

        keterangan = []
        if jk: keterangan.append(f"JK: {jk.group(1).strip()}")
        if gol_darah: keterangan.append(f"Gol: {gol_darah.group(1).strip()}")
        if agama: keterangan.append(f"Agama: {agama.group(1).strip()}")
        if status: keterangan.append(f"Status: {status.group(2).strip()}")
        if pekerjaan: keterangan.append(f"Pekerjaan: {pekerjaan.group(1).strip()}")
        if kel_desa: keterangan.append(f"Kel/Desa: {kel_desa.group(2).strip()}")
        if kecamatan: keterangan.append(f"Kecamatan: {kecamatan.group(2).strip()}")
        if kewarganegaraan: keterangan.append(f"WN: {kewarganegaraan.group(1).strip()}")
        if berlaku: keterangan.append(f"Berlaku: {berlaku.group(1).strip()}")
        if rt_rw: keterangan.append(f"RT/RW: {rt_rw.group(1).strip()}")

        return {
            "NIK / Nomor": nik.group(1).strip() if nik else "",
            "Nama": nama.group(1).strip() if nama else "",
            "Tempat/Tanggal Lahir": ttl.group(3).strip() if ttl else "",
            "Alamat": alamat.group(1).strip() if alamat else "",
            "Keterangan Tambahan": "; ".join(keterangan)
        }

    elif doc_type == "SIM":
        nomor = re.search(r"NO[:\s]*([A-Z0-9]+)", text)
        nama = re.search(r"NAMA[:\s]*([A-Z\s]+)", text)
        ttl = re.search(r"TEMPAT/TGL\s*LAHIR[:\s]*([A-Z0-9\s/-]+)", text)
        alamat = re.search(r"ALAMAT[:\s]*([A-Z0-9\s.,/]+)", text)

        return {
            "NIK / Nomor": nomor.group(1) if nomor else "",
            "Nama": nama.group(1).strip() if nama else "",
            "Tempat/Tanggal Lahir": ttl.group(1).strip() if ttl else "",
            "Alamat": alamat.group(1).strip() if alamat else "",
            "Keterangan Tambahan": "Surat Izin Mengemudi (SIM)"
        }

    elif doc_type == "NPWP":
        nomor = re.search(r"(\d{2}\.\d{3}\.\d{3}\.\d{1}-\d{3}\.\d{3})", text)
        nama = re.search(r"NAMA[:\s]*([A-Z\s]+)", text)
        alamat = re.search(r"ALAMAT[:\s]*([A-Z0-9\s.,/]+)", text)

        return {
            "NIK / Nomor": nomor.group(1) if nomor else "",
            "Nama": nama.group(1).strip() if nama else "",
            "Tempat/Tanggal Lahir": "",
            "Alamat": alamat.group(1).strip() if alamat else "",
            "Keterangan Tambahan": "NPWP"
        }

    else:
        return {
            "NIK / Nomor": "",
            "Nama": "",
            "Tempat/Tanggal Lahir": "",
            "Alamat": "",
            "Keterangan Tambahan": "Jenis dokumen tidak dikenali"
        }

# === VALIDASI DATA ===
def validate_row(row):
    errors = []
    if row["Jenis Dokumen"] == "KTP":
        if not re.fullmatch(r"\d{16}", str(row["NIK / Nomor"])):
            errors.append("NIK tidak valid (harus 16 digit angka)")
    elif row["Jenis Dokumen"] == "SIM":
        if len(str(row["NIK / Nomor"])) < 8:
            errors.append("Nomor SIM terlalu pendek")
    elif row["Jenis Dokumen"] == "NPWP":
        if not re.search(r"\d{2}\.\d{3}\.\d{3}\.\d{1}-\d{3}\.\d{3}", str(row["NIK / Nomor"])):
            errors.append("Format NPWP tidak standar")

    if not row["Nama"]:
        errors.append("Nama kosong")
    if row["Jenis Dokumen"] == "KTP" and not row["Alamat"]:
        errors.append("Alamat kosong")
    return errors

def highlight_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    col_index = {cell.value: cell.column for cell in ws[1]}
    for row in range(2, ws.max_row + 1):
        nik = ws.cell(row, col_index.get("NIK / Nomor")).value
        nama = ws.cell(row, col_index.get("Nama")).value
        alamat = ws.cell(row, col_index.get("Alamat")).value
        jenis = ws.cell(row, col_index.get("Jenis Dokumen")).value
        valid = True
        if not nama or (nik and len(str(nik)) < 8):
            valid = False
        if jenis == "KTP" and not alamat:
            valid = False
        fill = green_fill if valid else red_fill
        for cell in ws[row]:
            cell.fill = fill
    wb.save(file_path)

# === UPLOAD & PROSES ===
st.subheader("📸 Ambil / Upload Dokumen (KTP, SIM, NPWP)")

col1, col2 = st.columns(2)
with col1:
    img_camera = st.camera_input("Gunakan kamera (1 dokumen)")
with col2:
    img_files = st.file_uploader("📁 Upload banyak dokumen sekaligus", type=["jpg", "jpeg", "png"], accept_multiple_files=True)

images_to_process = []
if img_camera:
    images_to_process.append(("kamera_foto.png", Image.open(img_camera)))
if img_files:
    for file in img_files:
        images_to_process.append((file.name, Image.open(file)))

if images_to_process:
    batch_id = f"Batch_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    st.info(f"📦 Batch baru: **{batch_id}** – {len(images_to_process)} file sedang diproses...")
    results = []
    progress = st.progress(0)

    for i, (name, image) in enumerate(images_to_process):
        with st.spinner(f"🔍 Menganalisis {name}..."):
            processed = preprocess_image(auto_rotate_and_crop(image))
            text = pytesseract.image_to_string(processed, lang='eng+ind')
            doc_type = detect_document_type(text)
            fields = parse_document_fields(text, doc_type)
            row = {
                "Batch ID": batch_id,
                "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Nama File": name,
                "Jenis Dokumen": doc_type,
                **fields,
                "Teks Asli": text.strip()
            }
            results.append(row)
        progress.progress((i + 1) / len(images_to_process))

    append_data(results)

    # Validasi otomatis
    df_check = pd.DataFrame(results)
    invalids = []
    for _, row in df_check.iterrows():
        errs = validate_row(row)
        if errs:
            invalids.append((row["Nama File"], errs))

    if invalids:
        st.warning(f"⚠️ {len(invalids)} dokumen memiliki data tidak valid:")
        for file, errs in invalids:
            st.write(f"📄 {file}:")
            for e in errs:
                st.write(f"   • {e}")
    else:
        st.success("✅ Semua data valid!")

    highlight_excel(EXCEL_FILE)
    st.success(f"✅ {len(images_to_process)} dokumen berhasil diproses dan disimpan!")

# === TAMPILKAN DATA ===
st.subheader("📊 Data Dokumen Tersimpan")
df = load_data()
if len(df) == 0:
    st.info("Belum ada data dokumen yang tersimpan.")
else:
    filter_doc = st.selectbox("Filter Jenis Dokumen:", ["(Semua)"] + sorted(df["Jenis Dokumen"].unique().tolist()))
    if filter_doc != "(Semua)":
        df = df[df["Jenis Dokumen"] == filter_doc]
    st.dataframe(df, use_container_width=True, height=400)
    edited_df = st.data_editor(df, use_container_width=True, key="edit_data")
    if st.button("💾 Simpan Perubahan"):
        save_data(edited_df)
        st.success("✅ Data berhasil disimpan!")

    with open(EXCEL_FILE, "rb") as f:
        st.download_button(
            label="📥 Download Semua Data Excel",
            data=f,
            file_name=EXCEL_FILE,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
